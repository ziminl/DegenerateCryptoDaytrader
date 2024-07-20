## Functions that take information from buy and sell orders then log it to a spreadsheet
from openpyxl import Workbook, load_workbook
from resources import creds
from datetime import datetime

workbook_file_path = creds.trade_log_path


# If you manually create a workbook in Excel or in my case LibreOffice Calc beforehand, you'll have issues with openpyxl
def create_initial_workbook():
    workbook = Workbook()
    sheet = workbook.active
    headers = ["Order Number", "Date and Time", "Order Type", "Asset", "Bought Price", "Asset Amount Bought",
               "$ Amount Bought", "Sold Price", "Amount of Asset Sold", "$ Amount Sold", "Profit/Loss %",
               "Profit/Loss $ Amount", "$ Total Losses", "$ Total Profits", "$ Total Net"]
    sheet.append(headers)
    workbook.save(workbook_file_path)


def calculate_totals():
    workbook = load_workbook(workbook_file_path)

    sheet = workbook.active
    total_losses = 0
    total_profits = 0

    # Get the profit or loss of each row and calculate the total losses, total profits, and net
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=12, max_col=12, values_only=True):
        try:
            profit_loss = row[0]
            if profit_loss < 0:
                total_losses += profit_loss
            elif profit_loss > 0:
                total_profits += profit_loss
        except Exception as e:
            print(f"Incorrect data type in calculate_totals from logging_functions: {e}")
    total_net = total_profits + total_losses

    # Write the totals in the last three columns
    sheet.cell(row=2, column=13, value=round(total_losses, 2))
    sheet.cell(row=2, column=14, value=round(total_profits, 2))
    sheet.cell(row=2, column=15, value=round(total_net, 2))

    try:
        workbook.save(workbook_file_path)
        workbook.close()
    except Exception as e:
        print(f"Error in log_trade trying to save to workbook: {e}")


# trade and append it to the trade_log spreadsheet
def log_trade(*args):
    # Load in the workbook and select the main sheet then find the current free row
    workbook = None
    try:
        workbook = load_workbook(workbook_file_path)
    except FileNotFoundError:
        print("Workbook file does not seem to exist. Would you like to make a new workbook?")
        answer = input("YES/no: ")
        if answer == 'YES':
            create_initial_workbook()
            workbook = load_workbook(workbook_file_path)
        else:
            print("Skipping logging the trade!")
            return
    sheet = workbook.active
    next_row = sheet.max_row + 1
    order_number = sheet.max_row
    # Create a list of values to be appended from arguments passed to the function
    values_to_append = list(args)
    values_to_append = [order_number] + values_to_append

    # Add each value to their cells
    for column, value in enumerate(values_to_append, start=1):
        try:
            value = float(value)
        except:
            pass
        sheet.cell(row=next_row, column=column, value=value)
    #        print(f"Writing to row: {next_row}, column: {column}, value: {value}")
    # Save the changes to the spreadsheet
    try:
        workbook.save(workbook_file_path)
        workbook.close()
        print(f"Order appended to {workbook_file_path}")
    except Exception as e:
        print(f"Error in log_trade trying to save to workbook: {e}")


# Print a one-liner that the user can copy and paste the next time they want to run this trade as a one-liner
def repeat_one_liner(selected_user_options):
    # Make this a config setting that lets the user save what they have their alias to the log file
    program_alias = "degenerate_crypto_daytrader"
    command_option_list = []
    print("\n===========================\nCurrent options one-liner:\n===========================")
    # Add the --option value strings to a list then join the list
    for option, value in selected_user_options.items():
        command_option_list.append(f"--{option} {value}")
    # Take out the menu choice to be last so the callback function doesn't start the menu before it read the inputs
    menu_choice = command_option_list.pop()
    # Reverse the user inputs so click doesn't do a callback if the option dependency was defined after the parent
    command_option_list.reverse()
    command_option_string = ' '.join([string for string in command_option_list])
    current_one_liner = f'python3 {program_alias}.py {command_option_string} {menu_choice}'
    print(current_one_liner)
    # Get a timestamp and write the command to the history log file
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open('./resources/dcd_command_history.txt', 'a') as history:
        history.write(f'\n{timestamp} : {current_one_liner}\n')
    history.close()
    # Add in a ring buffer size to the history log file
    # Man page entry for this function


# Read the command history log file
def read_history(ctx, param, value):
    if value == "view":
        with open('resources/dcd_command_history.txt', 'r') as history:
            print(history.read())
            history.close()
            exit()
    elif value == "clear":
        # Open the file with write, which clears the contents of the file
        with open('./resources/dcd_command_history.txt', 'w') as history:
            history.close()
            exit()


if __name__ == "__main__":
    create_initial_workbook()
    calculate_totals()
    pass
